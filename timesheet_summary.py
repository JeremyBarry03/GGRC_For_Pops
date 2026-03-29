import argparse
import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


SUMMARY_HEADERS = ("Coach", "Position-AgeGrp", "Count")
WEEK_WORDS = {
    "one": 1,
    "two": 2,
    "three": 3,
    "four": 4,
    "five": 5,
    "six": 6,
    "seven": 7,
    "eight": 8,
}


def normalize_name(name, name_corrections: dict[str, str]) -> str | None:
    if not isinstance(name, str):
        return None

    cleaned = name.strip()
    if not cleaned:
        return None

    key = cleaned.lower()
    if key in {"xxx", "-"}:
        return None

    return name_corrections.get(key, cleaned)


def get_age_group(category_value) -> str:
    if isinstance(category_value, str) and "youth" in category_value.lower():
        return "Youth"
    return "Adult"


def get_role(role_value) -> str | None:
    if not isinstance(role_value, str):
        return None

    cleaned = role_value.strip()
    if not cleaned:
        return None

    lowered = cleaned.lower()
    if lowered.startswith("head"):
        return "Head"
    if lowered.startswith("asst") or lowered.startswith("assistant"):
        return "Asst"
    return None


def looks_like_summary_sheet(rows: list[list]) -> bool:
    if not rows:
        return False

    first_row = tuple("" if value is None else str(value).strip() for value in rows[0][:3])
    return first_row == SUMMARY_HEADERS


def extract_week_number(value: str) -> int | None:
    lowered = value.lower()
    for word, number in WEEK_WORDS.items():
        token = f"week {word}"
        if token in lowered:
            return number
    return None


def read_rows(path: Path) -> list[list]:
    ext = path.suffix.lower()

    if ext in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
        return rows

    if ext == ".csv":
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            return [row for row in csv.reader(handle)]

    if ext == ".xls":
        raise ValueError("`.xls` is not supported by this script. Save the file as `.xlsx` or `.csv` first.")

    raise ValueError(f"Unsupported file type: {ext}")


def build_summary(
    rows: list[list],
    name_corrections: dict[str, str],
    exclude_names: set[str],
    included_weeks: set[int] | None = None,
) -> list[dict[str, object]]:
    if looks_like_summary_sheet(rows):
        raise ValueError(
            "The input already looks like a summary workbook. Use the original timesheet file, not the generated summary."
        )

    counts: Counter[tuple[str, str]] = Counter()
    excluded_coaches: dict[str, str] = {}

    current_category = ""
    current_week: int | None = None

    for row in rows:
        if len(row) < 3:
            continue

        category = row[0]
        role = row[1]
        coach_values = row[2:9] if len(row) >= 9 else row[2:]

        category_text = category.strip() if isinstance(category, str) else ""
        if category_text.startswith("Session "):
            current_week = extract_week_number(category_text)
            current_category = ""
            continue

        if included_weeks is not None and current_week not in included_weeks:
            continue

        if category_text in {"Date", "Time", "PRIVATE LESSON"}:
            continue

        if category_text:
            current_category = category_text
        elif not current_category:
            continue

        role_clean = get_role(role)
        if role_clean is None:
            continue

        age_group = get_age_group(current_category)
        position_age_group = f"{role_clean}-{age_group}"

        for coach_raw in coach_values:
            coach = normalize_name(coach_raw, name_corrections)
            if coach is None:
                continue

            coach_key = coach.lower()
            if coach_key in exclude_names:
                excluded_coaches.setdefault(coach_key, coach)
                continue

            counts[(coach, position_age_group)] += 1

    summary_rows = [
        {"Coach": coach, "Position-AgeGrp": pos_age, "Count": count}
        for (coach, pos_age), count in counts.items()
    ]

    summary_rows.extend(
        {"Coach": coach, "Position-AgeGrp": "EXCLUDE", "Count": None}
        for _, coach in sorted(excluded_coaches.items(), key=lambda item: item[1].lower())
    )

    summary_rows.sort(key=lambda item: (str(item["Coach"]).lower(), str(item["Position-AgeGrp"]).lower()))
    return summary_rows


def autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = max_length + 2


def write_summary(output_path: Path, summary_rows: list[dict[str, object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"

    sheet.append(SUMMARY_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in summary_rows:
        sheet.append([row["Coach"], row["Position-AgeGrp"], row["Count"]])

    sheet.freeze_panes = "A2"
    autosize_columns(sheet)
    workbook.save(output_path)


def print_summary(summary_rows: list[dict[str, object]]) -> None:
    headers = list(SUMMARY_HEADERS)
    table_rows = [[row["Coach"], row["Position-AgeGrp"], "" if row["Count"] is None else row["Count"]] for row in summary_rows]
    widths = []

    for index, header in enumerate(headers):
        values = [header] + [str(row[index]) for row in table_rows]
        widths.append(max(len(value) for value in values))

    def format_row(values: list[object]) -> str:
        return "  ".join(str(value).ljust(widths[index]) for index, value in enumerate(values))

    print(format_row(headers))
    print(format_row(["-" * width for width in widths]))
    for row in table_rows:
        print(format_row(row))


def parse_name_corrections(items: list[str]) -> dict[str, str]:
    parsed: dict[str, str] = {}
    for item in items:
        if "=" not in item:
            continue
        bad, good = item.split("=", 1)
        bad = bad.strip().lower()
        good = good.strip()
        if bad and good:
            parsed[bad] = good
    return parsed


def parse_weeks(items: list[str] | None) -> set[int] | None:
    if not items:
        return None

    weeks: set[int] = set()
    for item in items:
        item = item.strip()
        if not item:
            continue
        weeks.add(int(item))
    return weeks or None


def main() -> None:
    parser = argparse.ArgumentParser(description="Create coach summary from a timesheet workbook or CSV.")
    parser.add_argument("input", type=Path, help="Path to the original timesheet `.xlsx` or `.csv` file")
    parser.add_argument(
        "output",
        type=Path,
        nargs="?",
        default=Path("coach_summary.xlsx"),
        help="Path to the output Excel workbook",
    )
    parser.add_argument("--exclude", nargs="*", default=[], help="Optional coach names to mark as EXCLUDE")
    parser.add_argument("--correct", nargs="*", default=["Caitln=Caitlin"], help="Name corrections in bad=good format")
    parser.add_argument("--weeks", nargs="*", help="Optional week numbers to include, for example `--weeks 1 2`")
    args = parser.parse_args()

    input_path = args.input.resolve()
    output_path = args.output.resolve()

    if input_path == output_path:
        raise ValueError("Input and output cannot be the same file. Point `output` to a different workbook.")

    name_corrections = parse_name_corrections(args.correct)
    exclude_names = {name.strip().lower() for name in args.exclude if name.strip()}
    included_weeks = parse_weeks(args.weeks)

    rows = read_rows(input_path)
    summary_rows = build_summary(
        rows,
        name_corrections=name_corrections,
        exclude_names=exclude_names,
        included_weeks=included_weeks,
    )

    write_summary(output_path, summary_rows)
    print_summary(summary_rows)
    print(f"\nWrote summary to {output_path}")


if __name__ == "__main__":
    main()
